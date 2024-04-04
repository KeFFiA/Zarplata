import re

import openpyxl as ox
from openpyxl.styles import Font, Border, Side, NamedStyle, PatternFill

from DataBase import get_ante, get_color


def identify(path, path_success1):
    book = ox.load_workbook(f"{path}")
    sheet = book.active
    file_name = path_success1
    znach(sheet, book, file_name)


def znach(sheet, book, file_name):
    ante = get_ante()
    col = ["D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"]
    time = []
    time11 = []
    e = ''
    r = ''

    for row in range(2, sheet.max_row + 1):
        name = sheet[row][0].value
        if name in ante:
            sheet[f"W{row}"] = ante[name]
        for column in col:
            q = sheet[f"{column}{row}"].value
            time.append(q)
        for i in time:
            try:
                e += i
                for i1 in e:
                    if i1 == ':':
                        i1 = '+'
                        r += i1
                    else:
                        r += i1
                r = re.sub(pattern=r'[+][0]', repl=r"+0o", string=r)
                r = re.sub(pattern=r'[+](0o9)|[+](0o8)', repl=r"+0o7", string=str(r))
                time11.append(r)
                r = ''
                e = ''
            except:
                pass
        t = list(map(lambda x: x + '/60', time11))
        time.clear()
        time11.clear()
        time_itog = []
        j = 0.0
        for i in range(len(t)):
            time_itog.append(round(eval(t[i]), 2))
            j += round(eval(t[i]), 2)
        sheet[f"U{row}"] = j
    cells_names(sheet, book, file_name)


def cells_names(sheet, book, file_name):
    name_4_cells = {
        "U": "Часы",
        "V": "Общие часы",
        "W": "Ставка",
        "X": "Итого наличка",
        "Y": "На карту",
        "Z": "Остаток",
        "AA": "Оплачено"
    }
    for column in name_4_cells:
        sheet[f"{column}{1}"] = name_4_cells[column]
    formulas(sheet, book, file_name)


def formulas(sheet, book, file_name):
    for row in range(2, sheet.max_row + 1):
        try:
            sheet[f"V{row}"] = f"=U{row}"
            sheet[f"X{row}"] = f"=V{row}*W{row}"
            sheet[f"Z{row}"] = f"=X{row}-Y{row}"
        except AttributeError:
            pass
    styles(sheet, book, file_name)


def styles(sheet, book, file_name):
    colors = get_color()

    colums = ['V', 'W', 'X', 'Y', 'Z', 'AA']
    borders_sheet = NamedStyle(name='border')
    borders_sheet.border = Border(left=Side(border_style='thin'),
                                  top=Side(border_style='thin'),
                                  bottom=Side(border_style='thin'),
                                  right=Side(border_style='thin'))

    for column in range(20, 26):
        sheet[1][column].font = Font(bold=True)

    for row in range(1, sheet.max_row + 1):
        for column in range(20, 27):
            sheet[row][column].style = borders_sheet

    a = 0
    b = 0
    for row in range(1, sheet.max_row + 1):
        name = sheet[row + a][0].value
        stroka1 = sheet[row + 1 + a][0].value
        stroka2 = sheet[row + 2 + a][0].value
        if stroka1 == name and stroka2 == name:
            a += 1
            if name is not None:
                for i in colums:
                    sheet.merge_cells(f'{i}{row + a - 1}:{i}{row + 2 + a - 1}')
                    sheet[f"V{row + a - 1}"] = f"=U{row + b}+U{row + a}+U{row + 2 + a - 1}"
                b += 1
        elif stroka1 == name or stroka2 == name:
            a += 1
            if name is not None:
                for i in colums:
                    sheet.merge_cells(f'{i}{row + a - 1}:{i}{row + 1 + a - 1}')
                    sheet[f"V{row + a - 1}"] = f"=U{row + b}+U{row + 1 + a - 1}"
                b += 1

    for row in range(2, sheet.max_row + 1):
        name = sheet[row][0].value
        for column in range(0, 27):
            if name in colors:
                sheet[row][column].fill = PatternFill(patternType='solid', start_color=f'{colors[name]}')

    sheet.freeze_panes = "B2"

    book.save(filename=file_name)
